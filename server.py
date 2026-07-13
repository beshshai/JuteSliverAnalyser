import os
import sys

# In a windowed (console=False) PyInstaller build there is no console, so
# sys.stdout / sys.stderr are None. Any print() or library writing to them
# raises AttributeError and silently kills the app before it can start the
# server. Give them a harmless place to write instead.
if getattr(sys, 'frozen', False) and (sys.stdout is None or sys.stderr is None):
    import io
    sys.stdout = sys.stdout or io.StringIO()
    sys.stderr = sys.stderr or io.StringIO()

import uuid
import json
import base64
import math
import struct
import zlib
from pathlib import Path
from flask import Flask, request, jsonify, send_file, send_from_directory, redirect, make_response
from werkzeug.utils import secure_filename
import urllib.request
import urllib.error
import psycopg2
from psycopg2.extras import RealDictCursor

# ── paths ──────────────────────────────────────────────────────────────────────
if getattr(sys, 'frozen', False):
    BASE_DIR = Path(sys.executable).parent
    # PyInstaller one-file builds unpack bundled data (like 'public') into a
    # temp dir at runtime (sys._MEIPASS), not next to the exe.
    BUNDLE_DIR = Path(getattr(sys, '_MEIPASS', BASE_DIR))
else:
    BASE_DIR = Path(__file__).parent
    BUNDLE_DIR = BASE_DIR

PUBLIC_DIR = BUNDLE_DIR / 'public'

app = Flask(__name__, static_folder=str(PUBLIC_DIR))

# ── database (Supabase Postgres) ────────────────────────────────────────────────
# Set DATABASE_URL in your environment to the connection string from
# Supabase → Project Settings → Database → Connection string (URI).
DATABASE_URL = os.environ.get('DATABASE_URL')
if not DATABASE_URL:
    raise RuntimeError(
        'DATABASE_URL is not set. Add it in your Render environment variables '
        '(Supabase → Project Settings → Database → Connection string).'
    )

class _DBWrapper:
    """Thin shim so the rest of the file can keep using sqlite-style '?'
    placeholders and conn.execute(...).fetchone()/.fetchall() calls."""
    def __init__(self, conn):
        self._conn = conn
    def execute(self, sql, params=()):
        cur = self._conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(sql.replace('?', '%s'), params)
        return cur
    def commit(self):
        self._conn.commit()
    def __enter__(self):
        return self
    def __exit__(self, exc_type, exc, tb):
        if exc_type is None:
            try:
                self._conn.commit()
            except Exception:
                pass
        self._conn.close()

def get_db():
    conn = psycopg2.connect(DATABASE_URL)
    return _DBWrapper(conn)

# ── file storage (Supabase Storage) ─────────────────────────────────────────────
# Set these in your environment too:
#   SUPABASE_URL          e.g. https://xxxxxxxx.supabase.co
#   SUPABASE_SERVICE_KEY   the "service_role" key (Project Settings → API)
#   SUPABASE_BUCKET        storage bucket name, e.g. "jute-uploads" (make it Public)
SUPABASE_URL = (os.environ.get('SUPABASE_URL') or '').rstrip('/')
SUPABASE_SERVICE_KEY = os.environ.get('SUPABASE_SERVICE_KEY')
SUPABASE_BUCKET = os.environ.get('SUPABASE_BUCKET', 'jute-uploads')

def _storage_configured():
    return bool(SUPABASE_URL and SUPABASE_SERVICE_KEY)

def storage_upload(key, data, content_type='image/jpeg'):
    """Uploads bytes to Supabase Storage and returns the public URL."""
    if not _storage_configured():
        raise RuntimeError('SUPABASE_URL / SUPABASE_SERVICE_KEY are not set.')
    url = f'{SUPABASE_URL}/storage/v1/object/{SUPABASE_BUCKET}/{key}'
    req = urllib.request.Request(url, data=data, method='POST', headers={
        'Authorization': f'Bearer {SUPABASE_SERVICE_KEY}',
        'Content-Type': content_type,
        'x-upsert': 'true',
    })
    with urllib.request.urlopen(req, timeout=30) as resp:
        resp.read()
    return f'{SUPABASE_URL}/storage/v1/object/public/{SUPABASE_BUCKET}/{key}'

def storage_delete(key):
    if not _storage_configured() or not key:
        return
    url = f'{SUPABASE_URL}/storage/v1/object/{SUPABASE_BUCKET}/{key}'
    req = urllib.request.Request(url, method='DELETE', headers={
        'Authorization': f'Bearer {SUPABASE_SERVICE_KEY}',
    })
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            resp.read()
    except urllib.error.HTTPError:
        pass  # already gone / never existed — fine to ignore on delete

def storage_key_from_url(image_path):
    """Recovers the storage object key from a public URL we generated."""
    prefix = f'{SUPABASE_URL}/storage/v1/object/public/{SUPABASE_BUCKET}/'
    if image_path and image_path.startswith(prefix):
        return image_path[len(prefix):]
    return None

DEFAULT_MACHINE_TYPES = [
    'Spreader', 'Inter-Spreader', 'Breaker-Card', 'Finisher-Card(rolls)',
    'Drawhead(sliver)', 'Draw-1', 'Draw-2', 'Draw-3', 'Spinning'
]

def init_db():
    with get_db() as conn:
        cur = conn._conn.cursor()
        cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL UNIQUE,
            created_at TEXT NOT NULL DEFAULT to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS')
        );
        CREATE TABLE IF NOT EXISTS batches (
            id TEXT PRIMARY KEY,
            user_id INTEGER NOT NULL REFERENCES users(id),
            name TEXT NOT NULL,
            notes TEXT,
            closed INTEGER NOT NULL DEFAULT 0,
            closed_at TEXT,
            machine_type TEXT,
            created_at TEXT NOT NULL DEFAULT to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS')
        );
        CREATE TABLE IF NOT EXISTS samples (
            id SERIAL PRIMARY KEY,
            user_id INTEGER NOT NULL REFERENCES users(id),
            batch_id TEXT REFERENCES batches(id) ON DELETE SET NULL,
            original_filename TEXT,
            image_path TEXT NOT NULL,
            width INTEGER,
            height INTEGER,
            score INTEGER,
            mean_angle_deg REAL,
            resultant_length_r REAL,
            circular_variance REAL,
            angular_stddev_deg REAL,
            edge_pixel_count INTEGER,
            histogram_json TEXT,
            notes TEXT,
            machine_type TEXT,
            analyzed INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL DEFAULT to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS')
        );
        CREATE INDEX IF NOT EXISTS idx_samples_user ON samples(user_id);
        CREATE INDEX IF NOT EXISTS idx_samples_created ON samples(created_at);
        CREATE INDEX IF NOT EXISTS idx_samples_batch ON samples(batch_id);
        CREATE INDEX IF NOT EXISTS idx_samples_machine ON samples(machine_type);

        CREATE TABLE IF NOT EXISTS machines (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL UNIQUE,
            is_default INTEGER NOT NULL DEFAULT 0,
            sort_order INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL DEFAULT to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS')
        );

        CREATE TABLE IF NOT EXISTS width_samples (
            id SERIAL PRIMARY KEY,
            user_id INTEGER NOT NULL REFERENCES users(id),
            set_label TEXT NOT NULL DEFAULT 'A',
            original_filename TEXT,
            image_path TEXT NOT NULL,
            image_width INTEGER,
            image_height INTEGER,
            rows_measured INTEGER,
            mean_width_px REAL,
            median_width_px REAL,
            min_width_px REAL,
            max_width_px REAL,
            sd_width_px REAL,
            row_widths_json TEXT,
            notes TEXT,
            created_at TEXT NOT NULL DEFAULT to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS')
        );
        CREATE INDEX IF NOT EXISTS idx_width_samples_user ON width_samples(user_id);
        CREATE INDEX IF NOT EXISTS idx_width_samples_set ON width_samples(set_label);
        """)
        conn._conn.commit()

        # ── seed the fixed default machine types (once) ──────────────────────
        for i, name in enumerate(DEFAULT_MACHINE_TYPES):
            cur.execute(
                "INSERT INTO machines (name, is_default, sort_order) VALUES (%s, 1, %s) "
                "ON CONFLICT (name) DO NOTHING",
                (name, i)
            )
        conn._conn.commit()

init_db()

# ── image processing (pure Python, no Pillow needed for basic ops) ─────────────
def decode_image_to_rgba(file_bytes, filename):
    """Decode image bytes to (rgba_bytes, width, height) using only stdlib."""
    ext = Path(filename).suffix.lower()
    if ext in ('.jpg', '.jpeg'):
        return decode_jpeg(file_bytes)
    elif ext == '.png':
        return decode_png(file_bytes)
    else:
        raise ValueError(f"Unsupported format: {ext}")

def decode_jpeg(data):
    """Minimal JPEG decoder using pillow if available, else raise helpful error."""
    try:
        from PIL import Image
        import io
        img = Image.open(io.BytesIO(data)).convert('RGBA')
        w, h = img.size
        return bytes(img.tobytes()), w, h
    except ImportError:
        raise RuntimeError("Pillow not available")

def decode_png(data):
    try:
        from PIL import Image
        import io
        img = Image.open(io.BytesIO(data)).convert('RGBA')
        w, h = img.size
        return bytes(img.tobytes()), w, h
    except ImportError:
        raise RuntimeError("Pillow not available")

def resize_rgba(rgba, w, h, target_w, target_h):
    """Nearest-neighbour resize — fast, no deps."""
    out = bytearray(target_w * target_h * 4)
    x_ratio = w / target_w
    y_ratio = h / target_h
    for ty in range(target_h):
        for tx in range(target_w):
            sx = int(tx * x_ratio)
            sy = int(ty * y_ratio)
            si = (sy * w + sx) * 4
            di = (ty * target_w + tx) * 4
            out[di:di+4] = rgba[si:si+4]
    return bytes(out)

def jpeg_bytes(rgba, w, h):
    """Encode RGBA to JPEG bytes in memory (for uploading to Supabase Storage)."""
    from PIL import Image
    import io
    img = Image.frombytes('RGBA', (w, h), rgba).convert('RGB')
    buf = io.BytesIO()
    img.save(buf, 'JPEG', quality=82)
    return buf.getvalue()

# ── fiber analysis ─────────────────────────────────────────────────────────────
def to_gray(rgba, w, h):
    gray = []
    for i in range(0, len(rgba), 4):
        r, g, b = rgba[i], rgba[i+1], rgba[i+2]
        gray.append(0.299 * r + 0.587 * g + 0.114 * b)
    return gray

def sobel_orientations(gray, w, h, mag_threshold=12):
    angles, mags = [], []
    for y in range(1, h - 1):
        for x in range(1, w - 1):
            i = y * w + x
            tl, tc, tr = gray[i-w-1], gray[i-w], gray[i-w+1]
            ml, mr = gray[i-1], gray[i+1]
            bl, bc, br = gray[i+w-1], gray[i+w], gray[i+w+1]
            sx = (tr + 2*mr + br) - (tl + 2*ml + bl)
            sy = (bl + 2*bc + br) - (tl + 2*tc + tr)
            mag = math.sqrt(sx*sx + sy*sy)
            if mag > mag_threshold:
                ang = math.atan2(sy, sx) * 180 / math.pi
                ang = (ang + 180) % 180
                fiber_ang = (ang + 90) % 180
                angles.append(fiber_ang)
                mags.append(mag)
    return angles, mags

def weighted_circular_stats(angles, mags):
    sum_w = sum_sin = sum_cos = 0
    for a, w in zip(angles, mags):
        rad = a * 2 * math.pi / 180
        sum_sin += w * math.sin(rad)
        sum_cos += w * math.cos(rad)
        sum_w += w
    sum_w = sum_w or 1
    R = math.sqrt(sum_sin**2 + sum_cos**2) / sum_w
    mean_angle = math.atan2(sum_sin, sum_cos) * 180 / math.pi / 2
    mean_angle = (mean_angle % 180 + 180) % 180
    circ_var = 1 - R
    log_val = -2 * math.log(max(R + 1e-9, 1e-9))
    circ_std = min(math.sqrt(max(log_val, 0)) * 180 / math.pi, 90)
    return mean_angle, R, circ_var, circ_std

def build_histogram(angles, mags, bins=36):
    hist = [0.0] * bins
    for a, m in zip(angles, mags):
        b = min(int(a / (180 / bins)), bins - 1)
        hist[b] += m
    max_val = max(hist) or 1
    return [v / max_val for v in hist]

def analyze_fiber(rgba, w, h):
    if len(angles := sobel_orientations(g := to_gray(rgba, w, h), w, h)[0]) < 20:
        return {'error': 'Not enough texture detected.'}
    angles, mags = sobel_orientations(g, w, h)
    if len(angles) < 20:
        return {'error': 'Not enough texture detected.'}
    mean_angle, R, circ_var, circ_std = weighted_circular_stats(angles, mags)
    hist = build_histogram(angles, mags)
    score = max(0, min(100, round((1 - circ_var) * 100)))
    return {
        'width': w, 'height': h, 'score': score,
        'meanAngleDeg': round(mean_angle, 1),
        'resultantLengthR': round(R, 2),
        'circularVariance': round(circ_var, 2),
        'angularStdDevDeg': round(circ_std, 1),
        'edgePixelCount': len(angles),
        'histogram': hist,
    }

# ── strand-width (mass variation) analysis ─────────────────────────────────────
def otsu_threshold(gray):
    """Classic Otsu's method on a 0-255 grayscale histogram. Returns a threshold
    that best separates two populations (strand vs background)."""
    hist = [0] * 256
    for v in gray:
        hist[min(255, max(0, int(v)))] += 1
    total = len(gray)
    if total == 0:
        return 127
    sum_all = sum(i * c for i, c in enumerate(hist))
    sum_b = 0.0
    weight_b = 0
    max_var = 0.0
    threshold = 127
    for t in range(256):
        weight_b += hist[t]
        if weight_b == 0:
            continue
        weight_f = total - weight_b
        if weight_f == 0:
            break
        sum_b += t * hist[t]
        mean_b = sum_b / weight_b
        mean_f = (sum_all - sum_b) / weight_f
        between_var = weight_b * weight_f * (mean_b - mean_f) ** 2
        if between_var > max_var:
            max_var = between_var
            threshold = t
    return threshold

def measure_strand_width(rgba, w, h):
    """Measures the strand's horizontal width (in pixels) on each row of the
    image, then returns per-row widths plus summary stats.

    Approach: convert to grayscale, find a separating threshold with Otsu's
    method, decide whether the strand is the darker or lighter region (the
    minority of pixels overall is assumed to be the strand, since the
    background usually fills more of the frame), then for each row take the
    widest contiguous run of "strand" pixels as that row's width. Rows with
    no detected strand pixels are skipped (e.g. blank margin at top/bottom).
    """
    gray = to_gray(rgba, w, h)
    threshold = otsu_threshold(gray)

    below = sum(1 for v in gray if v < threshold)
    above = len(gray) - below
    # Assume the strand is the minority class (background usually dominates
    # the frame). This also self-corrects for light-strand-on-dark-bg vs
    # dark-strand-on-light-bg photos.
    strand_is_dark = below <= above

    row_widths = []
    row_centers = []
    for y in range(h):
        base = y * w
        run_start = None
        best_len = 0
        best_start = 0
        x = 0
        while x < w:
            v = gray[base + x]
            is_strand = (v < threshold) if strand_is_dark else (v >= threshold)
            if is_strand:
                if run_start is None:
                    run_start = x
            else:
                if run_start is not None:
                    run_len = x - run_start
                    if run_len > best_len:
                        best_len = run_len
                        best_start = run_start
                    run_start = None
            x += 1
        if run_start is not None:
            run_len = w - run_start
            if run_len > best_len:
                best_len = run_len
                best_start = run_start
        if best_len > 0:
            row_widths.append(best_len)
            row_centers.append(best_start + best_len / 2)

    if len(row_widths) < max(5, h * 0.05):
        return {'error': 'Could not detect a clear strand in this image. Try a photo with more contrast between the sliver and the background.'}

    n = len(row_widths)
    mean_w = sum(row_widths) / n
    sorted_w = sorted(row_widths)
    median_w = sorted_w[n // 2] if n % 2 == 1 else (sorted_w[n // 2 - 1] + sorted_w[n // 2]) / 2
    var_w = sum((v - mean_w) ** 2 for v in row_widths) / max(1, n - 1)
    sd_w = math.sqrt(var_w)

    return {
        'imageWidth': w, 'imageHeight': h,
        'rowsMeasured': n,
        'meanWidthPx': round(mean_w, 2),
        'medianWidthPx': round(median_w, 2),
        'minWidthPx': min(row_widths),
        'maxWidthPx': max(row_widths),
        'sdWidthPx': round(sd_w, 2),
        'rowWidths': row_widths,
    }

# ── AI validation ──────────────────────────────────────────────────────────────
def validate_jute_image(image_bytes, mime_type):
    api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        return {'valid': True}
    b64 = base64.b64encode(image_bytes).decode()
    payload = json.dumps({
        'model': 'claude-haiku-4-5-20251001',
        'max_tokens': 200,
        'messages': [{
            'role': 'user',
            'content': [
                {'type': 'image', 'source': {'type': 'base64', 'media_type': mime_type, 'data': b64}},
                {'type': 'text', 'text': 'You are a quality-control gate for a jute fiber analysis app. Examine this image and decide if it shows jute sliver, jute fiber, raw jute, or similar fibrous textile material suitable for parallelization analysis.\n\nReply with ONLY a JSON object, no other text:\n{"valid": true, "reason": "brief reason"}\nor\n{"valid": false, "reason": "brief reason explaining what the image actually shows"}\n\nBe strict: textbooks, people, food, landscapes, documents, random objects, or anything not clearly showing fibrous/textile material should be rejected.'}
            ]
        }]
    }).encode()
    req = urllib.request.Request(
        'https://api.anthropic.com/v1/messages',
        data=payload,
        headers={'Content-Type': 'application/json', 'x-api-key': api_key, 'anthropic-version': '2023-06-01'}
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read())
            text = ''.join(b.get('text', '') for b in data.get('content', [])).strip()
            text = text.replace('```json', '').replace('```', '').strip()
            return json.loads(text)
    except Exception:
        return {'valid': True}

# ── session helpers ────────────────────────────────────────────────────────────
def get_session_user(request):
    user_id = request.cookies.get('user_id')
    if not user_id:
        return None
    with get_db() as conn:
        row = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
        return dict(row) if row else None

def require_user(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        user = get_session_user(request)
        if not user:
            return jsonify({'error': 'Not logged in'}), 401
        request.user = user
        return f(*args, **kwargs)
    return decorated

# ── session routes ─────────────────────────────────────────────────────────────
@app.post('/api/session')
def create_session():
    name = (request.json or {}).get('name', '').strip()
    if not name: return jsonify({'error': 'Name is required'}), 400
    if len(name) > 50: return jsonify({'error': 'Name too long'}), 400
    with get_db() as conn:
        row = conn.execute('SELECT * FROM users WHERE name = ?', (name,)).fetchone()
        if not row:
            conn.execute('INSERT INTO users (name) VALUES (?)', (name,))
            conn.commit()
            row = conn.execute('SELECT * FROM users WHERE name = ?', (name,)).fetchone()
        user = dict(row)
    resp = make_response(jsonify({'user': user}))
    resp.set_cookie('user_id', str(user['id']), max_age=365*24*3600, samesite='Lax', httponly=True)
    return resp

@app.get('/api/session')
def get_session():
    user = get_session_user(request)
    if not user: return jsonify({'error': 'Not logged in'}), 401
    return jsonify({'user': user})

@app.post('/api/logout')
def logout():
    resp = make_response(jsonify({'ok': True}))
    resp.delete_cookie('user_id')
    return resp

@app.post('/api/quit')
def quit_app():
    """Shut down the server cleanly from the UI."""
    import threading, os, signal
    threading.Timer(0.3, lambda: os.kill(os.getpid(), signal.SIGTERM)).start()
    return jsonify({'ok': True})

# ── batch routes ───────────────────────────────────────────────────────────────
@app.post('/api/batches')
@require_user
def create_batch():
    name = (request.json or {}).get('name', '').strip()
    if not name: return jsonify({'error': 'Batch name is required'}), 400
    notes = (request.json or {}).get('notes', '').strip() or None
    machine_type = (request.json or {}).get('machine_type', '').strip() or None
    batch_id = str(uuid.uuid4())
    with get_db() as conn:
        conn.execute('INSERT INTO batches (id, user_id, name, notes, machine_type) VALUES (?,?,?,?,?)',
                     (batch_id, request.user['id'], name, notes, machine_type))
        conn.commit()
        batch = conn.execute('''
            SELECT batches.*, users.name AS user_name,
                (SELECT COUNT(*) FROM samples WHERE batch_id = batches.id) AS sample_count
            FROM batches JOIN users ON users.id = batches.user_id WHERE batches.id = ?
        ''', (batch_id,)).fetchone()
    return jsonify({'batch': dict(batch)})

@app.get('/api/batches')
def list_batches():
    user_id = request.args.get('user_id')
    closed = request.args.get('closed')
    query = '''SELECT batches.*, users.name AS user_name,
        (SELECT COUNT(*) FROM samples WHERE batch_id = batches.id) AS sample_count,
        (SELECT ROUND(AVG(score),1) FROM samples WHERE batch_id = batches.id) AS avg_score
        FROM batches JOIN users ON users.id = batches.user_id WHERE 1=1'''
    params = []
    if user_id: query += ' AND batches.user_id = ?'; params.append(user_id)
    if closed is not None: query += ' AND batches.closed = ?'; params.append(int(closed))
    query += ' ORDER BY batches.closed ASC, batches.created_at DESC'
    with get_db() as conn:
        rows = conn.execute(query, params).fetchall()
    return jsonify({'batches': [dict(r) for r in rows]})

@app.patch('/api/batches/<batch_id>/close')
@require_user
def close_batch(batch_id):
    with get_db() as conn:
        batch = conn.execute('SELECT * FROM batches WHERE id = ?', (batch_id,)).fetchone()
        if not batch: return jsonify({'error': 'Batch not found'}), 404
        if batch['user_id'] != request.user['id']: return jsonify({'error': 'Not your batch'}), 403
        conn.execute("UPDATE batches SET closed=1, closed_at=to_char(NOW(),'YYYY-MM-DD HH24:MI:SS') WHERE id=?", (batch_id,))
        conn.commit()
    return jsonify({'ok': True})

@app.patch('/api/batches/<batch_id>/reopen')
@require_user
def reopen_batch(batch_id):
    with get_db() as conn:
        batch = conn.execute('SELECT * FROM batches WHERE id = ?', (batch_id,)).fetchone()
        if not batch: return jsonify({'error': 'Batch not found'}), 404
        if batch['user_id'] != request.user['id']: return jsonify({'error': 'Not your batch'}), 403
        conn.execute('UPDATE batches SET closed=0, closed_at=NULL WHERE id=?', (batch_id,))
        conn.commit()
    return jsonify({'ok': True})

@app.get('/api/batches/<batch_id>')
def get_batch(batch_id):
    with get_db() as conn:
        batch = conn.execute('''
            SELECT batches.*, users.name AS user_name,
                (SELECT COUNT(*) FROM samples WHERE batch_id = batches.id) AS sample_count,
                (SELECT ROUND(AVG(score),1) FROM samples WHERE batch_id = batches.id) AS avg_score
            FROM batches JOIN users ON users.id = batches.user_id WHERE batches.id = ?
        ''', (batch_id,)).fetchone()
        if not batch: return jsonify({'error': 'Not found'}), 404
        samples = conn.execute('''
            SELECT samples.*, users.name AS user_name FROM samples
            JOIN users ON users.id = samples.user_id
            WHERE samples.batch_id = ? ORDER BY samples.created_at ASC
        ''', (batch_id,)).fetchall()
    return jsonify({'batch': dict(batch), 'samples': [dict(s) for s in samples]})

# ── sample routes ──────────────────────────────────────────────────────────────
@app.post('/api/samples')
@require_user
def create_sample():
    if 'image' not in request.files:
        return jsonify({'error': 'No image uploaded'}), 400
    f = request.files['image']
    file_bytes = f.read()
    mime = f.mimetype or 'image/jpeg'

    validation = validate_jute_image(file_bytes, mime)
    if not validation.get('valid'):
        return jsonify({'error': 'This image does not appear to show jute fiber or sliver. ' + validation.get('reason', '')}), 422

    batch_id = (request.form.get('batch_id') or '').strip() or None

    MAX_DIM = 480
    try:
        rgba, w, h = decode_image_to_rgba(file_bytes, f.filename or 'image.jpg')
    except Exception as e:
        return jsonify({'error': f'Could not decode image: {e}'}), 422

    scale = min(1, MAX_DIM / max(w, h))
    tw, th = max(1, round(w * scale)), max(1, round(h * scale))
    if scale < 1:
        rgba = resize_rgba(rgba, w, h, tw, th)
        w, h = tw, th

    result = analyze_fiber(rgba, w, h)
    if 'error' in result:
        return jsonify({'error': result['error']}), 422

    filename = str(uuid.uuid4()) + '.jpg'
    key = f'{batch_id}/{filename}' if batch_id else filename
    image_path = storage_upload(key, jpeg_bytes(rgba, w, h))

    notes = (request.form.get('notes') or '').strip() or None
    machine_type = (request.form.get('machine_type') or '').strip() or None
    # If no machine_type given but a batch is set, inherit the batch's machine_type
    if not machine_type and batch_id:
        with get_db() as conn:
            brow = conn.execute('SELECT machine_type FROM batches WHERE id = ?', (batch_id,)).fetchone()
            if brow:
                machine_type = brow['machine_type']
    with get_db() as conn:
        cur = conn.execute('''
            INSERT INTO samples (user_id, batch_id, original_filename, image_path, width, height,
                score, mean_angle_deg, resultant_length_r, circular_variance,
                angular_stddev_deg, edge_pixel_count, histogram_json, notes, machine_type)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) RETURNING id
        ''', (request.user['id'], batch_id, f.filename, image_path, w, h,
              result['score'], result['meanAngleDeg'], result['resultantLengthR'],
              result['circularVariance'], result['angularStdDevDeg'],
              result['edgePixelCount'], json.dumps(result['histogram']), notes, machine_type))
        new_id = cur.fetchone()['id']
        conn.commit()
        row = conn.execute('''
            SELECT samples.*, users.name AS user_name FROM samples
            JOIN users ON users.id = samples.user_id WHERE samples.id = ?
        ''', (new_id,)).fetchone()
    return jsonify({'sample': dict(row)})

@app.get('/api/samples')
def list_samples():
    q = 'SELECT samples.*, users.name AS user_name FROM samples JOIN users ON users.id = samples.user_id WHERE 1=1'
    params = []
    for key, col in [('user_id', 'samples.user_id'), ('batch_id', 'samples.batch_id'), ('machine_type', 'samples.machine_type')]:
        if v := request.args.get(key): q += f' AND {col} = ?'; params.append(v)
    if v := request.args.get('from'): q += ' AND samples.created_at >= ?'; params.append(v)
    if v := request.args.get('to'): q += ' AND samples.created_at <= ?'; params.append(v + ' 23:59:59')
    q += ' ORDER BY samples.created_at DESC LIMIT ? OFFSET ?'
    params += [int(request.args.get('limit', 50)), int(request.args.get('offset', 0))]
    with get_db() as conn:
        rows = conn.execute(q, params).fetchall()
    return jsonify({'samples': [dict(r) for r in rows]})

@app.get('/api/samples/<int:sample_id>')
def get_sample(sample_id):
    with get_db() as conn:
        row = conn.execute('SELECT samples.*, users.name AS user_name FROM samples JOIN users ON users.id = samples.user_id WHERE samples.id = ?', (sample_id,)).fetchone()
    if not row: return jsonify({'error': 'Not found'}), 404
    return jsonify({'sample': dict(row)})

@app.delete('/api/samples/<int:sample_id>')
@require_user
def delete_sample(sample_id):
    with get_db() as conn:
        row = conn.execute('SELECT * FROM samples WHERE id = ?', (sample_id,)).fetchone()
        if not row: return jsonify({'error': 'Not found'}), 404
        if row['user_id'] != request.user['id']: return jsonify({'error': 'Not your sample'}), 403
        conn.execute('DELETE FROM samples WHERE id = ?', (sample_id,))
        conn.commit()
    storage_delete(storage_key_from_url(row['image_path']))
    return jsonify({'ok': True})

# ── width-sample routes (mass variation analysis, from photos) ─────────────────
@app.post('/api/width-samples')
@require_user
def create_width_sample():
    if 'image' not in request.files:
        return jsonify({'error': 'No image uploaded'}), 400
    f = request.files['image']
    file_bytes = f.read()
    mime = f.mimetype or 'image/jpeg'

    validation = validate_jute_image(file_bytes, mime)
    if not validation.get('valid'):
        return jsonify({'error': 'This image does not appear to show jute fiber or sliver. ' + validation.get('reason', '')}), 422

    set_label = (request.form.get('set_label') or 'A').strip().upper()
    if set_label not in ('A', 'B'):
        set_label = 'A'

    MAX_DIM = 900  # keep more resolution than fiber analysis since we need width precision
    try:
        rgba, w, h = decode_image_to_rgba(file_bytes, f.filename or 'image.jpg')
    except Exception as e:
        return jsonify({'error': f'Could not decode image: {e}'}), 422

    scale = min(1, MAX_DIM / max(w, h))
    tw, th = max(1, round(w * scale)), max(1, round(h * scale))
    if scale < 1:
        rgba = resize_rgba(rgba, w, h, tw, th)
        w, h = tw, th

    result = measure_strand_width(rgba, w, h)
    if 'error' in result:
        return jsonify({'error': result['error']}), 422

    filename = str(uuid.uuid4()) + '.jpg'
    image_path = storage_upload(f'width/{filename}', jpeg_bytes(rgba, w, h))

    notes = (request.form.get('notes') or '').strip() or None
    with get_db() as conn:
        cur = conn.execute('''
            INSERT INTO width_samples (user_id, set_label, original_filename, image_path,
                image_width, image_height, rows_measured, mean_width_px, median_width_px,
                min_width_px, max_width_px, sd_width_px, row_widths_json, notes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?) RETURNING id
        ''', (request.user['id'], set_label, f.filename, image_path, w, h,
              result['rowsMeasured'], result['meanWidthPx'], result['medianWidthPx'],
              result['minWidthPx'], result['maxWidthPx'], result['sdWidthPx'],
              json.dumps(result['rowWidths']), notes))
        new_id = cur.fetchone()['id']
        conn.commit()
        row = conn.execute('''
            SELECT width_samples.*, users.name AS user_name FROM width_samples
            JOIN users ON users.id = width_samples.user_id WHERE width_samples.id = ?
        ''', (new_id,)).fetchone()
    return jsonify({'sample': dict(row)})

@app.get('/api/width-samples')
def list_width_samples():
    q = '''SELECT width_samples.*, users.name AS user_name FROM width_samples
           JOIN users ON users.id = width_samples.user_id WHERE 1=1'''
    params = []
    if v := request.args.get('user_id'): q += ' AND width_samples.user_id = ?'; params.append(v)
    if v := request.args.get('set_label'): q += ' AND width_samples.set_label = ?'; params.append(v.upper())
    q += ' ORDER BY width_samples.created_at DESC LIMIT ? OFFSET ?'
    params += [int(request.args.get('limit', 100)), int(request.args.get('offset', 0))]
    with get_db() as conn:
        rows = conn.execute(q, params).fetchall()
    return jsonify({'samples': [dict(r) for r in rows]})

@app.delete('/api/width-samples/<int:sample_id>')
@require_user
def delete_width_sample(sample_id):
    with get_db() as conn:
        row = conn.execute('SELECT * FROM width_samples WHERE id = ?', (sample_id,)).fetchone()
        if not row: return jsonify({'error': 'Not found'}), 404
        if row['user_id'] != request.user['id']: return jsonify({'error': 'Not your sample'}), 403
        conn.execute('DELETE FROM width_samples WHERE id = ?', (sample_id,))
        conn.commit()
    storage_delete(storage_key_from_url(row['image_path']))
    return jsonify({'ok': True})

@app.get('/api/users')
def list_users():
    with get_db() as conn:
        rows = conn.execute('SELECT id, name FROM users ORDER BY name').fetchall()
    return jsonify({'users': [dict(r) for r in rows]})

# ── machine types ────────────────────────────────────────────────────────────
@app.get('/api/machines')
def list_machines():
    with get_db() as conn:
        rows = conn.execute(
            'SELECT id, name, is_default FROM machines ORDER BY is_default DESC, sort_order ASC, name ASC'
        ).fetchall()
    return jsonify({'machines': [dict(r) for r in rows]})

@app.post('/api/machines')
@require_user
def create_machine():
    name = ((request.json or {}).get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Machine name is required'}), 400
    if len(name) > 60:
        return jsonify({'error': 'Machine name too long'}), 400
    with get_db() as conn:
        existing = conn.execute('SELECT id, name, is_default FROM machines WHERE LOWER(name) = LOWER(?)', (name,)).fetchone()
        if existing:
            return jsonify({'machine': dict(existing)})
        max_order = conn.execute('SELECT COALESCE(MAX(sort_order), 0) AS m FROM machines').fetchone()['m']
        cur = conn.execute(
            'INSERT INTO machines (name, is_default, sort_order) VALUES (?, 0, ?) RETURNING id',
            (name, max_order + 1)
        )
        new_id = cur.fetchone()['id']
        conn.commit()
        row = conn.execute('SELECT id, name, is_default FROM machines WHERE id = ?', (new_id,)).fetchone()
    return jsonify({'machine': dict(row)})

@app.delete('/api/machines/<int:machine_id>')
@require_user
def delete_machine(machine_id):
    with get_db() as conn:
        row = conn.execute('SELECT * FROM machines WHERE id = ?', (machine_id,)).fetchone()
        if not row:
            return jsonify({'error': 'Machine not found'}), 404
        if row['is_default']:
            return jsonify({'error': 'Default machine types cannot be removed'}), 400
        conn.execute('DELETE FROM machines WHERE id = ?', (machine_id,))
        conn.commit()
    return jsonify({'ok': True})

# ── stats ──────────────────────────────────────────────────────────────────────
@app.get('/api/stats/summary')
def stats_summary():
    group_by = request.args.get('group_by', 'day').lower()
    machine_filter = (request.args.get('machine_type') or request.args.get('machine_name') or '').strip() or None
    with get_db() as conn:
        per_user_q = '''
            SELECT users.id AS user_id, users.name AS user_name, COUNT(*) AS sample_count,
                ROUND(AVG(samples.score), 1) AS avg_score
            FROM samples JOIN users ON users.id = samples.user_id
        '''
        per_user_params = []
        if machine_filter:
            per_user_q += ' WHERE samples.machine_type = ?'
            per_user_params.append(machine_filter)
        per_user_q += ' GROUP BY users.id ORDER BY avg_score DESC'
        per_user = conn.execute(per_user_q, per_user_params).fetchall()

        formats = {'hour': '%Y-%m-%d %H:00', 'day': '%Y-%m-%d', 'month': '%Y-%m', 'year': '%Y'}
        if group_by == 'log':
            q = 'SELECT id, created_at AS day, score AS avg_score, 1 AS sample_count, machine_type AS machine FROM samples'
            params = []
            if machine_filter:
                q += ' WHERE machine_type = ?'; params.append(machine_filter)
            q += ' ORDER BY created_at ASC'
            over_time = conn.execute(q, params).fetchall()
        elif group_by == 'batch':
            q = '''
                SELECT batches.id AS batch_id, batches.name AS day,
                    ROUND(AVG(samples.score), 1) AS avg_score, COUNT(samples.id) AS sample_count
                FROM batches JOIN samples ON samples.batch_id = batches.id
            '''
            params = []
            if machine_filter:
                q += ' WHERE samples.machine_type = ?'; params.append(machine_filter)
            q += ' GROUP BY batches.id HAVING COUNT(samples.id) > 0 ORDER BY batches.created_at ASC'
            over_time = conn.execute(q, params).fetchall()
        elif group_by == 'batch_log':
            batch_id = request.args.get('batch_id', '').strip()
            if not batch_id: return jsonify({'error': 'batch_id required'}), 400
            q = 'SELECT id, created_at AS day, score AS avg_score, 1 AS sample_count FROM samples WHERE batch_id = ?'
            params = [batch_id]
            if machine_filter:
                q += ' AND machine_type = ?'; params.append(machine_filter)
            q += ' ORDER BY created_at ASC'
            over_time = conn.execute(q, params).fetchall()
        elif group_by == 'machine':
            q = '''
                SELECT COALESCE(samples.machine_type, 'Unspecified') AS day,
                    ROUND(AVG(samples.score), 1) AS avg_score, COUNT(samples.id) AS sample_count,
                    MIN(samples.created_at) AS first_used
                FROM samples
                GROUP BY day ORDER BY avg_score DESC
            '''
            over_time = conn.execute(q).fetchall()
        elif group_by == 'machine_log':
            machine_id_filter = request.args.get('machine_name', '').strip()
            if not machine_id_filter: return jsonify({'error': 'machine_name required'}), 400
            over_time = conn.execute(
                'SELECT id, created_at AS day, score AS avg_score, 1 AS sample_count FROM samples WHERE machine_type = ? ORDER BY created_at ASC',
                (machine_id_filter,)
            ).fetchall()
        else:
            fmt = formats.get(group_by, formats['day'])
            q = "SELECT strftime(?, created_at) AS day, ROUND(AVG(score),1) AS avg_score, COUNT(*) AS sample_count FROM samples"
            params = [fmt]
            if machine_filter:
                q += ' WHERE machine_type = ?'; params.append(machine_filter)
            q += ' GROUP BY day ORDER BY day ASC'
            over_time = conn.execute(q, params).fetchall()

    return jsonify({'perUser': [dict(r) for r in per_user], 'overTime': [dict(r) for r in over_time], 'groupBy': group_by})

# ── reports ────────────────────────────────────────────────────────────────────
HIST_BINS = 36

def combine_histograms(samples):
    combined = [0.0] * HIST_BINS
    for s in samples:
        if not s.get('histogram_json'): continue
        try: hist = json.loads(s['histogram_json'])
        except: continue
        weight = s.get('edge_pixel_count') or 1
        for i in range(min(HIST_BINS, len(hist))):
            combined[i] += hist[i] * weight
    max_val = max(combined) or 1
    return [v / max_val for v in combined]

def build_report_stats(samples):
    n = len(samples)
    if not n:
        return {'sampleCount': 0, 'avgScore': None, 'minScore': None, 'maxScore': None,
                'avgMeanAngle': None, 'avgCircularVariance': None, 'avgAngularStdDev': None,
                'scoreDistribution': {'good': 0, 'moderate': 0, 'poor': 0},
                'histogram': [0.0] * HIST_BINS}
    scores = [s['score'] for s in samples]
    dist = {'good': sum(1 for s in scores if s >= 75), 'moderate': sum(1 for s in scores if 50 <= s < 75), 'poor': sum(1 for s in scores if s < 50)}
    return {
        'sampleCount': n,
        'avgScore': round(sum(scores) / n, 1),
        'minScore': min(scores), 'maxScore': max(scores),
        'avgMeanAngle': round(sum(s['mean_angle_deg'] for s in samples) / n, 1),
        'avgCircularVariance': round(sum(s['circular_variance'] for s in samples) / n, 2),
        'avgAngularStdDev': round(sum(s['angular_stddev_deg'] for s in samples) / n, 1),
        'scoreDistribution': dist,
        'histogram': combine_histograms(samples),
    }

@app.get('/api/reports/user/<int:user_id>')
def report_user(user_id):
    with get_db() as conn:
        user = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
        if not user: return jsonify({'error': 'User not found'}), 404
        samples = conn.execute('SELECT samples.*, users.name AS user_name FROM samples JOIN users ON users.id = samples.user_id WHERE samples.user_id = ? ORDER BY samples.created_at DESC', (user_id,)).fetchall()
    samples = [dict(s) for s in samples]
    return jsonify({'subject': {'type': 'user', 'id': dict(user)['id'], 'name': dict(user)['name']}, 'stats': build_report_stats(samples), 'samples': samples})

@app.get('/api/reports/batch/<batch_id>')
def report_batch(batch_id):
    with get_db() as conn:
        batch = conn.execute('SELECT batches.*, users.name AS user_name FROM batches JOIN users ON users.id = batches.user_id WHERE batches.id = ?', (batch_id,)).fetchone()
        if not batch: return jsonify({'error': 'Batch not found'}), 404
        samples = conn.execute('SELECT samples.*, users.name AS user_name FROM samples JOIN users ON users.id = samples.user_id WHERE samples.batch_id = ? ORDER BY samples.created_at DESC', (batch_id,)).fetchall()
    batch, samples = dict(batch), [dict(s) for s in samples]
    return jsonify({'subject': {'type': 'batch', 'id': batch['id'], 'name': batch['name'], 'notes': batch.get('notes'), 'ownerName': batch['user_name']}, 'stats': build_report_stats(samples), 'samples': samples})

# ── excel export ───────────────────────────────────────────────────────────────
def make_excel_report(subject, stats, samples):
    """Build and return an openpyxl Workbook for the given report data."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    DARK = "1D1D1B"
    GREEN = "1D9E75"
    AMBER = "BA7517"
    RED   = "E24B4A"
    HEAD_FILL = PatternFill("solid", fgColor="2B2B29")
    SUB_FILL  = PatternFill("solid", fgColor="F0EDE6")
    thin = Side(style="thin", color="D0CEC6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hcell(ws, row, col, value, bold=True, color="FFFFFF", fill=None, align="left", num_fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Arial", bold=bold, color=color, size=10)
        c.alignment = Alignment(horizontal=align, vertical="center")
        if fill: c.fill = fill
        if num_fmt: c.number_format = num_fmt
        c.border = border
        return c

    def dcell(ws, row, col, value, bold=False, align="left", num_fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Arial", bold=bold, size=10, color=DARK)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border = border
        if num_fmt: c.number_format = num_fmt
        return c

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    subject_type = subject.get("type", "")
    title_cell.value = f"{'Batch' if subject_type == 'batch' else 'Individual'} Report — {subject.get('name', '')}"
    title_cell.font = Font(name="Arial", bold=True, size=14, color=DARK)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    if subject_type == "batch" and subject.get("ownerName"):
        ws.merge_cells("A2:H2")
        ws["A2"].value = f"Owner: {subject['ownerName']}"
        ws["A2"].font = Font(name="Arial", size=10, color="6B6A64")
    if subject.get("notes"):
        ws.merge_cells("A3:H3")
        ws["A3"].value = f"Notes: {subject['notes']}"
        ws["A3"].font = Font(name="Arial", size=10, italic=True, color="6B6A64")

    # Summary stats block
    r = 5
    headers = ["Metric", "Value"]
    for ci, h in enumerate(headers, 1):
        hcell(ws, r, ci, h, fill=HEAD_FILL)
    ws.row_dimensions[r].height = 18
    r += 1

    summary_rows = [
        ("Sample count",         stats.get("sampleCount")),
        ("Average score",        stats.get("avgScore")),
        ("Min score",            stats.get("minScore")),
        ("Max score",            stats.get("maxScore")),
        ("Avg mean angle (°)",   stats.get("avgMeanAngle")),
        ("Avg angular std dev (°)", stats.get("avgAngularStdDev")),
        ("Avg circular variance",stats.get("avgCircularVariance")),
        ("Good samples (≥75)",   stats.get("scoreDistribution", {}).get("good")),
        ("Moderate samples (50–74)", stats.get("scoreDistribution", {}).get("moderate")),
        ("Poor samples (<50)",   stats.get("scoreDistribution", {}).get("poor")),
    ]
    for label, val in summary_rows:
        dcell(ws, r, 1, label, bold=True)
        dcell(ws, r, 2, val, align="right")
        ws.row_dimensions[r].height = 16
        r += 1

    # ODF section header
    r += 1
    ws.merge_cells(f"A{r}:H{r}")
    odf_title = ws[f"A{r}"]
    odf_title.value = "Orientation Distribution Function (ODF) — Histogram Data"
    odf_title.font = Font(name="Arial", bold=True, size=11, color=DARK)
    odf_title.fill = SUB_FILL
    ws.row_dimensions[r].height = 20
    r += 1

    hist = stats.get("histogram") or []
    bin_size = 180.0 / len(hist) if hist else 5
    odf_headers = ["Bin #", "Angle Start (°)", "Angle End (°)", "Normalized Weight", "Interpretation"]
    for ci, h in enumerate(odf_headers, 1):
        hcell(ws, r, ci, h, fill=HEAD_FILL)
    ws.row_dimensions[r].height = 18
    r += 1

    for i, v in enumerate(hist):
        angle_start = round(i * bin_size, 1)
        angle_end   = round((i + 1) * bin_size, 1)
        if v >= 0.8:   interp = "Dominant"
        elif v >= 0.5: interp = "Strong"
        elif v >= 0.2: interp = "Moderate"
        else:           interp = "Weak"
        dcell(ws, r, 1, i + 1, align="center")
        dcell(ws, r, 2, angle_start, align="right", num_fmt="0.0")
        dcell(ws, r, 3, angle_end,   align="right", num_fmt="0.0")
        dcell(ws, r, 4, round(v, 2),  align="right", num_fmt="0.00")
        dcell(ws, r, 5, interp)
        ws.row_dimensions[r].height = 15
        r += 1

    # ODF key metrics
    r += 1
    ws.merge_cells(f"A{r}:H{r}")
    ws[f"A{r}"].value = "ODF Key Metrics"
    ws[f"A{r}"].font = Font(name="Arial", bold=True, size=11, color=DARK)
    ws[f"A{r}"].fill = SUB_FILL
    ws.row_dimensions[r].height = 20
    r += 1

    avg_angle = stats.get("avgMeanAngle") or 0
    avg_cv    = stats.get("avgCircularVariance") or 0
    avg_std   = stats.get("avgAngularStdDev") or 0
    avg_angle_md = abs(90.0 - avg_angle)
    orient_idx = round(1 - avg_cv, 2)

    # Compute Herman's f from histogram (distribution-weighted), not just mean angle.
    # Each bin's angle is converted from CD (image horizontal) to MD (90° − CD).
    # f = (3·<cos²φ_MD> − 1) / 2  where <cos²φ_MD> = Σ(Hᵢ·cos²φᵢ) / ΣHᵢ
    if hist and sum(hist) > 0:
        cos2_sum = 0.0
        w_sum    = 0.0
        for i, w in enumerate(hist):
            bin_angle_cd = (i + 0.5) * bin_size   # centre of bin, degrees from CD
            bin_angle_md = abs(90.0 - bin_angle_cd)
            cos2_sum += w * math.cos(math.radians(bin_angle_md)) ** 2
            w_sum    += w
        cos2_phi = cos2_sum / w_sum
        hermans_f = round((3 * cos2_phi - 1) / 2, 2)
    else:
        # fallback to mean-angle method if no histogram
        hermans_f = round((3 * math.cos(math.radians(avg_angle_md))**2 - 1) / 2, 2)

    if hermans_f >= 0.9:   hf_qual = "Excellent — fibers very well parallelized"
    elif hermans_f >= 0.7: hf_qual = "Good — fibers well aligned"
    elif hermans_f >= 0.4: hf_qual = "Moderate alignment"
    else:                   hf_qual = "Poor — fibers disordered / tangled"

    odf_metrics = [
        ("Average Mean Fiber Angle — CD (°)", avg_angle, "Measured from cross direction (horizontal in image)"),
        ("Average Mean Fiber Angle — MD (°)", round(avg_angle_md, 1), "Angle from machine direction (90° − CD angle); used for Herman's f"),
        ("Average Angular Std Dev (°)",  avg_std,   "Spread of fiber angles; lower = better parallelism"),
        ("Average Circular Variance",     avg_cv,    "0 = perfectly aligned, 1 = random"),
        ("Orientation Index (R)",         orient_idx, "1 = perfect alignment, 0 = random"),
        ("Herman's Orientation Factor (f)", hermans_f, hf_qual),
    ]
    for label, val, desc in odf_metrics:
        dcell(ws, r, 1, label, bold=True)
        dcell(ws, r, 2, val, align="right")
        dcell(ws, r, 3, desc)
        ws.row_dimensions[r].height = 16
        r += 1

    # Column widths
    for col, width in [(1, 32), (2, 16), (3, 16), (4, 20), (5, 28)]:
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Sheet 2: Sample Data ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Sample Data")
    cols2 = [
        ("ID", 6), ("Uploader", 18), ("Original File", 22), ("Date", 18),
        ("Score", 8), ("Quality", 12), ("Mean Angle (°)", 16),
        ("Angular Std Dev (°)", 20), ("Circular Variance", 18),
        ("Resultant R", 14), ("Edge Pixels", 14), ("Notes", 30),
    ]
    for ci, (label, width) in enumerate(cols2, 1):
        hcell(ws2, 1, ci, label, fill=HEAD_FILL)
        ws2.column_dimensions[get_column_letter(ci)].width = width
    ws2.row_dimensions[1].height = 18

    for ri, s in enumerate(samples, 2):
        score = s.get("score") or 0
        if score >= 75:   quality = "Good"
        elif score >= 50: quality = "Moderate"
        else:              quality = "Poor"
        row_vals = [
            s.get("id"), s.get("user_name"), s.get("original_filename"),
            s.get("created_at"), score, quality,
            s.get("mean_angle_deg"), s.get("angular_stddev_deg"),
            s.get("circular_variance"), s.get("resultant_length_r"),
            s.get("edge_pixel_count"), s.get("notes") or "",
        ]
        for ci, v in enumerate(row_vals, 1):
            c = ws2.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Arial", size=10, color=DARK)
            c.border = border
            c.alignment = Alignment(vertical="center")
            if ci == 5:  # Score — color-code
                col_hex = "C8F7E4" if score >= 75 else ("FFF0C0" if score >= 50 else "FFD5D5")
                c.fill = PatternFill("solid", fgColor=col_hex)
                c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[ri].height = 15

    # ── Sheet 3: ODF per Sample ───────────────────────────────────────────────
    ws3 = wb.create_sheet("ODF Per Sample")
    ws3.merge_cells("A1:G1")
    ws3["A1"].value = "Orientation Distribution — Per Sample Histogram Values"
    ws3["A1"].font = Font(name="Arial", bold=True, size=12, color=DARK)
    ws3.row_dimensions[1].height = 22

    # Fixed angle headers
    if samples:
        first_hist = []
        for s in samples:
            if s.get("histogram_json"):
                try: first_hist = json.loads(s["histogram_json"]); break
                except: pass
        n_bins = len(first_hist) if first_hist else 36
        bin_sz = 180.0 / n_bins

        hcell(ws3, 2, 1, "Sample ID", fill=HEAD_FILL)
        hcell(ws3, 2, 2, "Score",     fill=HEAD_FILL)
        for bi in range(n_bins):
            hcell(ws3, 2, bi + 3, f"{round(bi*bin_sz,0):.0f}°–{round((bi+1)*bin_sz,0):.0f}°", fill=HEAD_FILL)
        ws3.column_dimensions["A"].width = 12
        ws3.column_dimensions["B"].width = 8

        for ri, s in enumerate(samples, 3):
            ws3.cell(row=ri, column=1, value=s.get("id")).font = Font(name="Arial", size=9)
            ws3.cell(row=ri, column=2, value=s.get("score")).font = Font(name="Arial", size=9)
            try: hist_vals = json.loads(s.get("histogram_json") or "[]")
            except: hist_vals = []
            for bi, v in enumerate(hist_vals):
                c = ws3.cell(row=ri, column=bi + 3, value=round(v, 2))
                c.font = Font(name="Arial", size=9)
                c.number_format = "0.00"
                intensity = int(v * 200)
                hex_color = f"{255 - intensity:02X}FF{255 - intensity:02X}"
                c.fill = PatternFill("solid", fgColor=hex_color)
            ws3.row_dimensions[ri].height = 14

    return wb

@app.get('/api/export/batch/<batch_id>')
def export_batch_excel(batch_id):
    with get_db() as conn:
        batch = conn.execute('SELECT batches.*, users.name AS user_name FROM batches JOIN users ON users.id = batches.user_id WHERE batches.id = ?', (batch_id,)).fetchone()
        if not batch: return jsonify({'error': 'Batch not found'}), 404
        samples = conn.execute('SELECT samples.*, users.name AS user_name FROM samples JOIN users ON users.id = samples.user_id WHERE samples.batch_id = ? ORDER BY samples.created_at DESC', (batch_id,)).fetchall()
    batch, samples = dict(batch), [dict(s) for s in samples]
    subject = {'type': 'batch', 'id': batch['id'], 'name': batch['name'], 'notes': batch.get('notes'), 'ownerName': batch['user_name']}
    stats = build_report_stats(samples)
    wb = make_excel_report(subject, stats, samples)
    if wb is None:
        return jsonify({'error': 'openpyxl not available'}), 500
    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = ''.join(c if c.isalnum() or c in '-_ ' else '_' for c in batch['name'])
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f"batch_{safe_name}.xlsx")

@app.get('/api/export/user/<int:user_id>')
def export_user_excel(user_id):
    with get_db() as conn:
        user = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
        if not user: return jsonify({'error': 'User not found'}), 404
        samples = conn.execute('SELECT samples.*, users.name AS user_name FROM samples JOIN users ON users.id = samples.user_id WHERE samples.user_id = ? ORDER BY samples.created_at DESC', (user_id,)).fetchall()
    user, samples = dict(user), [dict(s) for s in samples]
    subject = {'type': 'user', 'id': user['id'], 'name': user['name']}
    stats = build_report_stats(samples)
    wb = make_excel_report(subject, stats, samples)
    if wb is None:
        return jsonify({'error': 'openpyxl not available'}), 500
    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = ''.join(c if c.isalnum() or c in '-_ ' else '_' for c in user['name'])
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f"user_{safe_name}.xlsx")

# ── static / page routes ───────────────────────────────────────────────────────
@app.get('/')
def index():
    user = get_session_user(request)
    if user: return redirect('/dashboard')
    return send_from_directory(str(PUBLIC_DIR), 'login.html')

@app.get('/dashboard')
def dashboard():
    if not get_session_user(request): return redirect('/')
    return send_from_directory(str(PUBLIC_DIR), 'dashboard.html')

@app.get('/batches')
def batches_page():
    if not get_session_user(request): return redirect('/')
    return send_from_directory(str(PUBLIC_DIR), 'batches.html')

@app.get('/trends')
def trends():
    if not get_session_user(request): return redirect('/')
    return send_from_directory(str(PUBLIC_DIR), 'trends.html')

@app.get('/report')
def report():
    if not get_session_user(request): return redirect('/')
    return send_from_directory(str(PUBLIC_DIR), 'report.html')

@app.get('/mass-variation')
def mass_variation():
    if not get_session_user(request): return redirect('/')
    return send_from_directory(str(PUBLIC_DIR), 'mass_variation.html')

@app.get('/mass-report')
def mass_report():
    if not get_session_user(request): return redirect('/')
    return send_from_directory(str(PUBLIC_DIR), 'mass_report.html')

@app.get('/sliver-width')
def sliver_width():
    if not get_session_user(request): return redirect('/')
    return send_from_directory(str(PUBLIC_DIR), 'sliver_width.html')

@app.get('/width-analysis')
def width_analysis_redirect():
    # Old URL kept for backwards compatibility — page was renamed to
    # "Mass Variation Analysis" and moved to /mass-variation.
    return redirect('/mass-variation')

@app.route('/<path:filename>')
def serve_static(filename):
    return send_from_directory(str(PUBLIC_DIR), filename)

# ── main ───────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    import webbrowser, threading
    port = int(os.environ.get('PORT', 3000))
    print(f'Jute sliver analyzer running at http://localhost:{port}')
    threading.Timer(1.5, lambda: webbrowser.open(f'http://localhost:{port}')).start()
    app.run(host='0.0.0.0', port=port, debug=False)
